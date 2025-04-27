import streamlit as st
import pandas as pd
import dask.dataframe as dd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from io import BytesIO

# Hide main menu, footer, header
hide_streamlit_style = """
    <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    </style>
"""
st.markdown(hide_streamlit_style, unsafe_allow_html=True)

# Autofit function
def autofit_excel(writer):
    for sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]
        for col in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

# Preprocessing
def preprocess_dataframe(df):
    exclude = [
        'STAFF SOCIAL LOAN', 'STAFF VEHICLE LOAN', 'STAFF HOME LOAN',
        'STAFF FLEXIBLE LOAN', 'STAFF HOME LOAN(COF)', 'STAFF VEHICLE FACILITY LOAN (EVF)'
    ]
    df['Ac Type Desc'] = df['Ac Type Desc'].str.strip().str.upper()
    exclude = [x.upper() for x in exclude]
    df = df[~df['Ac Type Desc'].isin(exclude)]
    df = df[df['Limit'] != 0]
    df = df[~df['Main Code'].isin(['AcType Total', 'Grand Total'])]
    return df

# Branch comparison
def calculate_common_branchcode(sheets_1, sheets_2, writer):
    combined_df = pd.DataFrame()

    for df1 in sheets_1.values():
        for df2 in sheets_2.values():
            if all(c in df1.columns for c in ['Branch Code', 'Balance', 'Main Code', 'Limit']) and \
               all(c in df2.columns for c in ['Branch Code', 'Balance', 'Main Code', 'Limit']):

                df1 = preprocess_dataframe(df1.compute())
                df2 = preprocess_dataframe(df2.compute())

                df1['Branch Code'] = df1['Branch Code'].astype(str).str.zfill(3)
                df2['Branch Code'] = df2['Branch Code'].astype(str).str.zfill(3)

                df1_grouped = df1.groupby('Branch Code').agg({'Balance': 'sum', 'Branch Code': 'count'})
                df2_grouped = df2.groupby('Branch Code').agg({'Balance': 'sum', 'Branch Code': 'count'})

                df1_grouped.columns = ['Previous Balance Sum', 'Previous Count']
                df2_grouped.columns = ['New Balance Sum', 'New Count']

                combined_df = pd.merge(df1_grouped, df2_grouped, left_index=True, right_index=True, how='outer').fillna(0)
                combined_df['Change'] = combined_df['New Balance Sum'] - combined_df['Previous Balance Sum']
                combined_df['Percent Change'] = (combined_df['Change'] / combined_df['Previous Balance Sum'].replace(0, pd.NA)) * 100
                combined_df['Percent Change'] = combined_df['Percent Change'].fillna(0)

    if not combined_df.empty:
        combined_df = combined_df.round(2)
        combined_df.to_excel(writer, sheet_name='Branch', index=True)
        ws = writer.sheets['Branch']

        # Format Branch Code column as text
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
            for cell in row:
                cell.number_format = '@'

        # Format numbers with commas
        number_columns = [2, 4, 5]  # Previous Balance Sum, New Balance Sum, Change
        for col in number_columns:
            for row in ws.iter_rows(min_row=2, min_col=col, max_col=col, max_row=ws.max_row):
                for cell in row:
                    cell.number_format = '#,##0'

        # Format percentages
        percent_col = 6  # Percent Change
        for row in ws.iter_rows(min_row=2, min_col=percent_col, max_col=percent_col, max_row=ws.max_row):
            for cell in row:
                cell.number_format = '0.00%'

# (Other functions stay unchanged: preprocess_dataframe, autofit_excel, etc.)

# Main function
def main():
    st.title("Excel File Comparison Tool")

    previous_file = st.file_uploader("Upload Previous Period's Excel File", type=["xlsx"])
    current_file = st.file_uploader("Upload This Period's Excel File", type=["xlsx"])

    if previous_file and current_file:
        if st.button("Start Processing"):
            with st.spinner("Processing..."):
                try:
                    df_previous = pd.read_excel(previous_file)
                    df_this = pd.read_excel(current_file)

                    sheets_1 = {name: dd.from_pandas(sheet, npartitions=1) for name, sheet in pd.read_excel(previous_file, sheet_name=None).items()}
                    sheets_2 = {name: dd.from_pandas(sheet, npartitions=1) for name, sheet in pd.read_excel(current_file, sheet_name=None).items()}

                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        calculate_common_branchcode(sheets_1, sheets_2, writer)
                        autofit_excel(writer)

                    st.success("Processing Done!")
                    st.download_button(
                        label="Download Branch Comparison",
                        data=output.getvalue(),
                        file_name="branch_comparison.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except Exception as e:
                    st.error(f"Error: {e}")

if __name__ == "__main__":
    main()
